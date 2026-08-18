"""Prescription parsing and rule-AST compilation.

The workbook is read at runtime and is deliberately not redistributed. Every
source value is either compiled into a typed condition or recorded in the
``unresolved`` map; silent dropping is prohibited.
"""

from __future__ import annotations

import math
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from .models import Bound, Condition, Prescription

FIELD_MAP: dict[str, tuple[str, str | None, str]] = {
    "FDIIgnitionDay1": ("FFDI", None, "provisional"),
    "Temperature": ("temperature_c", "degC", "provisional"),
    "RelativeHumidity": ("relative_humidity_pct", "%", "provisional"),
    "WindSpeed": ("wind_speed_kmh", "km/h", "provisional"),
    "KBDI": ("KBDI", None, "provisional"),
    "FMCProfile": ("fmc_profile_pct", "%", "unmapped"),
    "FMCSurfaceInside": ("fmc_surface_inside_pct", "%", "unmapped"),
    "FMCSurfaceOutside": ("fmc_surface_outside_pct", "%", "unmapped"),
    "FMCNearSurface": ("fmc_near_surface_pct", "%", "unmapped"),
    "CuringMin": ("curing_pct", "%", "unmapped"),
    "DewPoint": ("dew_point_c", "degC", "unmapped"),
    "FMCBark": ("fmc_bark_pct", "%", "unmapped"),
    "FMCGullyControl": ("fmc_gully_control_pct", "%", "unmapped"),
    "FMCBetweenWindows": ("fmc_between_windows_pct", "%", "unmapped"),
    # Ground-wind height and conversion are not confirmed.
    "WindSpeedGround": ("wind_speed_ground_kmh", "km/h", "unmapped"),
}

METADATA_FIELDS = {
    "BurnClass",
    "PrescriptionOrder",
    "WindSpeedDescription",
    "DisplayCondition",
    "DisplaySpeciesComposition",
    "DisplayGrasslandsFuelHazard",
    "UseGFDI",
}

NULLS = {"", "null", "none", "nan", "n/a"}
RANGE_RE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*[-–]\s*(-?\d+(?:\.\d+)?)\s*$")
COMPARE_RE = re.compile(r"^\s*(<=|>=|<|>)\s*(-?\d+(?:\.\d+)?)\s*$")


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    return None if text.lower() in NULLS else text


def parse_threshold(
    value: Any,
    *,
    field: str,
    variable: str,
    unit: str | None,
    status: str,
    season: str | None = None,
) -> Condition:
    """Parse a range or inequality while preserving inclusive boundaries."""

    text = _clean(value)
    if text is None:
        raise ValueError("empty threshold")
    range_match = RANGE_RE.match(text)
    if range_match:
        lower, upper = (float(item) for item in range_match.groups())
        return Condition(
            field=field,
            variable=variable,
            unit=unit,
            lower=Bound(value=lower, inclusive=True),
            upper=Bound(value=upper, inclusive=True),
            season=season,
            source_text=text,
            operational_status=status,
        )
    compare_match = COMPARE_RE.match(text)
    if compare_match:
        operator, raw = compare_match.groups()
        bound = Bound(value=float(raw), inclusive="=" in operator)
        return Condition(
            field=field,
            variable=variable,
            unit=unit,
            lower=bound if operator.startswith(">") else None,
            upper=bound if operator.startswith("<") else None,
            season=season,
            source_text=text,
            operational_status=status,
        )
    raise ValueError(f"unsupported threshold syntax: {text!r}")


def parse_kbdi(value: Any) -> tuple[list[Condition], dict[str, str]]:
    """Compile explicit seasons; retain semantically unclear labels."""

    text = _clean(value)
    if text is None:
        return [], {}
    mapping = FIELD_MAP["KBDI"]
    if ":" not in text:
        return [
            parse_threshold(
                text,
                field="KBDI",
                variable=mapping[0],
                unit=mapping[1],
                status=mapping[2],
            )
        ], {}

    conditions: list[Condition] = []
    unresolved: dict[str, str] = {}
    for part in (piece.strip() for piece in text.split(",")):
        if ":" not in part:
            unresolved[f"KBDI:{len(unresolved)}"] = part
            continue
        label, threshold = (item.strip() for item in part.split(":", 1))
        season = label.lower()
        if season not in {"spring", "summer", "autumn", "winter"}:
            unresolved[f"KBDI:{label}"] = threshold
            continue
        conditions.append(
            parse_threshold(
                threshold,
                field="KBDI",
                variable=mapping[0],
                unit=mapping[1],
                status=mapping[2],
                season=season,
            )
        )
    return conditions, unresolved


def compile_record(record: dict[str, Any]) -> Prescription:
    burn_class = _clean(record.get("BurnClass"))
    if not burn_class:
        raise ValueError("record has no BurnClass")
    conditions: list[Condition] = []
    unresolved: dict[str, str] = {}
    metadata: dict[str, Any] = {}
    for field, raw in record.items():
        text = _clean(raw)
        if text is None:
            continue
        if field in METADATA_FIELDS:
            if field != "BurnClass":
                metadata[field] = raw
            continue
        if field == "KBDI":
            parsed, pending = parse_kbdi(raw)
            conditions.extend(parsed)
            unresolved.update(pending)
            continue
        if field in FIELD_MAP:
            variable, unit, status = FIELD_MAP[field]
            try:
                conditions.append(
                    parse_threshold(
                        raw,
                        field=field,
                        variable=variable,
                        unit=unit,
                        status=status,
                    )
                )
            except ValueError:
                unresolved[field] = text
            continue
        # Day 2/3 semantics and anomalous date values must not be guessed.
        unresolved[field] = text
    return Prescription(
        burn_class=burn_class,
        conditions=conditions,
        unresolved=unresolved,
        metadata=metadata,
    )


def compile_records(records: Iterable[dict[str, Any]]) -> list[Prescription]:
    prescriptions = [compile_record(record) for record in records]
    names = [item.burn_class for item in prescriptions]
    if len(names) != len(set(names)):
        raise ValueError("duplicate BurnClass values")
    return prescriptions


def load_prescriptions(path: str | Path) -> list[Prescription]:
    """Load all rows from the first workbook sheet without retaining the file."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency error path
        raise RuntimeError("openpyxl is required to read the prescriptions workbook") from exc
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records = [
        dict(zip(headers, row, strict=False))
        for row in rows
        if any(value is not None for value in row)
    ]
    result = compile_records(records)
    if len(result) != 43:
        raise ValueError(f"expected 43 burn classes, found {len(result)}")
    return result


def compilation_summary(prescriptions: list[Prescription]) -> dict[str, Any]:
    return {
        "burn_class_count": len(prescriptions),
        "condition_count": sum(len(item.conditions) for item in prescriptions),
        "mapped_condition_count": sum(
            condition.operational_status == "mapped"
            for item in prescriptions
            for condition in item.conditions
        ),
        "provisional_condition_count": sum(
            condition.operational_status == "provisional"
            for item in prescriptions
            for condition in item.conditions
        ),
        "unmapped_condition_count": sum(
            condition.operational_status == "unmapped"
            for item in prescriptions
            for condition in item.conditions
        ),
        "unresolved_value_count": sum(len(item.unresolved) for item in prescriptions),
        "classes_with_unresolved_values": sum(bool(item.unresolved) for item in prescriptions),
    }

